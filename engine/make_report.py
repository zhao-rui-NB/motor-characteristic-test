import os
import csv
import shutil
import time
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from engine.Motor import Motor
from engine.MultiScalePlotter import MultiScalePlotter


def make_report_a(motor: Motor, input_dir, output_path: str):
    template_path = 'report_template/電動機特性計算表a_20250728.xlsx'
    
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    try:
        shutil.copy(template_path, output_path)
    except PermissionError as e:
        print(f"[make_report_a] PermissionError: {e}")
        print(f"!!! Please close the file: {output_path} and try again.")
        return
    except FileNotFoundError as e:
        print(f"[make_report_a] FileNotFoundError: {e}")
        print(f"!!! Template file not found: {template_path}")
        return
    
        
    wb = load_workbook(output_path)  # 載入 Excel
    ws = wb.active                   # 使用第一個工作表
    
    ws['C4'] = motor.information_dict.get("試驗日期", "")
    ws['L4'] = motor.information_dict.get("印表日期", "")
    ws['C5'] = motor.information_dict.get("電腦編號", "")
    ws['H5'] = motor.information_dict.get("序號" , "")
    ws['M5'] = motor.information_dict.get("型式" , "")
    ws['C6'] = motor.information_dict.get("工作單號" , "")
    ws['H6'] = motor.information_dict.get("製造號碼" , "")
    ws['M6'] = motor.information_dict.get("廠牌" , "")
    
    ws['C7'] = motor.rated_voltage
    ws['G7'] = motor.rated_current
    ws['K7'] = motor.horsepower*0.746 if motor.horsepower else ''
    ws['M7'] = motor.horsepower
    
    ws['C8'] = motor.no_load_current
    ws['G8'] = motor.power_phases
    ws['K8'] = motor.frequency
    ws['O8'] = motor.poles
    
    ws['C9'] = motor.speed
    ws['G9'] = motor.information_dict.get("周圍溫度" , "")
    ws['K9'] = motor.information_dict.get("冷電阻" , "")
    ws['O9'] = motor.information_dict.get("熱電阻" , "")
    
    
    ws['C10'] = motor.information_dict.get("溫升" , "")
    ws['G10'] = motor.information_dict.get("絕緣種類" , "")
    
    ws['C11'] = motor.information_dict.get("定子規格" , "")
    ws['K11'] = motor.information_dict.get("轉子規格" , "")
    ws['C12'] = motor.information_dict.get("本線" , "")
    ws['K12'] = motor.information_dict.get("啟動線" , "")
    
    ws['C13'] = motor.information_dict.get("備註" , "")
    
    # ####################### data table #######################
    load_report_csv = os.path.join(input_dir, 'load_report_x.csv')
    if not os.path.exists(load_report_csv):
        print(f"[make_report_a] Load report CSV file not found: {load_report_csv}")
        return
    
    with open(load_report_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers_en = next(reader)
        headers_ch = next(reader)
        # print(f"headers_en: {headers_en}")
        # print(f"headers_ch: {headers_ch}")
        
        # each row last is the V header
        v_header = []
        csv_data = []
        for row in reader:
            v_header.append(row[-1].strip())
            csv_data.append(list(map(float, row[:-1])))
            
        # print("v_header:", v_header)
        # print("csv_data:", csv_data) 

    # b18
    table_start_row = 18
    table_start_col = 2
    
    for r, r_datas in enumerate(csv_data):
        now_row = table_start_row + r
        now_col = table_start_col
        for c, value in enumerate(r_datas):
            
            while True:
                cell = ws.cell(row=now_row, column=now_col)
                if isinstance(cell, MergedCell): # go to right cell
                    now_col += 1
                else:
                    break
            cell.value = value
            now_col += 1

    
    # excel_h_header = ['voltage','current', 'pf', 'speed', 'slip', 'torque', 'power', 'pout', 'eff' ]
    # excel_v_header = ['no_load', '25%', '50%', '75%', '100%', '125%', '150%', 'po_max', 'start', 'tq_max']
    
    # # try to match the header_en with the csv data
    # for c, hd1 in enumerate(headers_en):
    #     for r, hd2 in enumerate(v_header):
    #         # find hd1 in excel_h_header
    #         if hd1 in excel_h_header and hd2 in excel_v_header:
    #             # find the index of hd1 in excel_h_header
    #             h_index = excel_h_header.index(hd1)
    #             v_index = excel_v_header.index(hd2)
                
    #             cell = ws.cell(row=table_start_row + v_index, column=table_start_col + h_index)
    #             cell.value = csv_data[r][c]
    
    # save the workbook
    wb.save(output_path)
    wb.close()
    print(f"[make_report_a] Report saved to {output_path}")
    
def make_report_b(motor: Motor, input_dir, output_path: str):
    template_path = 'report_template/電動機特性計算表b1_20250728_rui.xlsx'

    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    try:
        shutil.copy(template_path, output_path)
    except PermissionError as e:
        print(f"[make_report_a] PermissionError: {e}")
        print(f"!!! Please close the file: {output_path} and try again.")
        return
    except FileNotFoundError as e:
        print(f"[make_report_a] FileNotFoundError: {e}")
        print(f"!!! Template file not found: {template_path}")
        return
    
    wb = load_workbook(output_path)  # 載入 Excel
    # ws = wb.active                   # 使用第一個工作表
    # open 工作表2
    ws = wb['工作表2']
    
    
    ws['C4'] = motor.information_dict.get("電腦編號", "")
    ws['L4'] = motor.information_dict.get("印表日期", "")
    
    
    ws['C5'] = motor.information_dict.get("型式" , "")
    ws['H5'] = motor.information_dict.get("製造號碼" , "")
    ws['L5'] = motor.information_dict.get("試驗日期", "")


    ws['A8'] = motor.horsepower
    ws['B8'] = motor.horsepower*0.746 if motor.horsepower else ''
    ws['C8'] = motor.power_phases
    ws['D8'] = motor.frequency
    ws['E8'] = motor.rated_voltage
    ws['G8'] = motor.rated_current
    ws['H8'] = motor.poles
    ws['I8'] = motor.speed
    ws['J8'] = motor.no_load_current
    ws['K8'] = motor.information_dict.get("溫升" , "")
    ws['L8'] = motor.information_dict.get("絕緣種類" , "")



    # ####################### data table #######################
    load_report_csv = os.path.join(input_dir, 'load_report_x.csv')
    if not os.path.exists(load_report_csv):
        print(f"[make_report_a] Load report CSV file not found: {load_report_csv}")
        return
    
    with open(load_report_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers_en = next(reader)
        headers_ch = next(reader)
        # print(f"headers_en: {headers_en}")
        # print(f"headers_ch: {headers_ch}")
        
        # each row last is the V header
        v_header = []
        csv_data = []
        for row in reader:
            v_header.append(row[-1].strip())
            csv_data.append(list(map(float, row[:-1])))
        
        if len(csv_data) > 7:
            # only keep the first 7 rows
            csv_data = csv_data[:7]
            
        # print("v_header:", v_header)
        # print("csv_data:", csv_data) 

    # b12
    table_start_row = 12
    table_start_col = 2
    
    for r, r_datas in enumerate(csv_data):
        now_row = table_start_row + r
        now_col = table_start_col
        for c, value in enumerate(r_datas):
            
            while True:
                cell = ws.cell(row=now_row, column=now_col)
                if isinstance(cell, MergedCell): # go to right cell
                    now_col += 1
                else:
                    break
            cell.value = value
            now_col += 1


    #### page 2 ####
    ws = wb['工作表1']  # 使用工作表1
    load_report_csv = os.path.join(input_dir, 'loadsort_x.csv')
    if not os.path.exists(load_report_csv):
        print(f"[make_report_a] Load report CSV file not found: {load_report_csv}")
        return

    with open(load_report_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        csv_data = []
        for row in reader:
            # remove front and end space
            new_row = [item.strip() for item in row]
            csv_data.append(new_row)
    
    # past all to 0,0
    for r, r_datas in enumerate(csv_data):
        for c, value in enumerate(r_datas):
            cell = ws.cell(row=r + 1, column=c + 1)
            # try to convert value to float, if fail, keep as string
            try:
                value = float(value)
            except:
                pass
            cell.value = value

    # save the workbook
    wb.save(output_path)
    wb.close()
    print(f"[make_report_b] Report saved to {output_path}")

def make_report_cns(motor: Motor, input_dir, output_path: str):
    template_path = 'report_template\電動機特性計算表_cns_20250728.xlsx'
    
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    try:
        shutil.copy(template_path, output_path)
    except PermissionError as e:
        print(f"[make_report_a] PermissionError: {e}")
        print(f"!!! Please close the file: {output_path} and try again.")
        return
    except FileNotFoundError as e:
        print(f"[make_report_a] FileNotFoundError: {e}")
        print(f"!!! Template file not found: {template_path}")
        return
    
    wb = load_workbook(output_path)  # 載入 Excel
    ws = wb.active                   # 使用第一個工作表
    
    
    ws['C4'] = motor.information_dict.get("電腦編號", "")
    ws['G4'] = motor.information_dict.get("序號" , "")
    ws['J4'] = motor.information_dict.get("型式" , "")
    ws['P4'] = motor.information_dict.get("試驗日期", "")
    
    ws['C5'] = motor.horsepower*0.746 if motor.horsepower else ''
    ws['E5'] = motor.horsepower
    ws['I5'] = motor.rated_current
    ws['M5'] = motor.no_load_current
    ws['Q5'] = motor.frequency
    
    ws['C6'] = motor.rated_voltage
    ws['F6'] = motor.poles
    ws['I6'] = motor.power_phases
    ws['M6'] = motor.speed
    ws['Q6'] = motor.information_dict.get("周圍溫度", "")
    


    # ####################### data table #######################
    load_report_csv = os.path.join(input_dir, 'cns14400_report.csv')
    if not os.path.exists(load_report_csv):
        print(f"[make_report_a] Load report CSV file not found: {load_report_csv}")
        return
    
    with open(load_report_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers_en = next(reader)
        # print(f"headers_en: {headers_en}")
        # print(f"headers_ch: {headers_ch}")
        
        # each row last is the V header
        v_index = []
        v_header = []
        csv_data = []
        for row in reader:
            v_index.append(row[0].strip())
            v_header.append(row[1].strip())
            csv_data.append(list(map(float, row[2:])))
        # print("csv_data:", csv_data) 

    # b18
    table_start_row = 9
    table_start_col = 7
    
    for r, r_datas in enumerate(csv_data):
        now_row = table_start_row + r
        now_col = table_start_col
        for c, value in enumerate(r_datas):
            
            while True:
                cell = ws.cell(row=now_row, column=now_col)
                if isinstance(cell, MergedCell): # go to right cell
                    now_col += 1
                else:
                    break
            cell.value = value
            now_col += 1

    # save the workbook
    wb.save(output_path)
    wb.close()
    print(f"[make_report_cns] Report saved to {output_path}")


def polt_motor_characteristic_fig(motor: Motor, input_dir, output_path: str):
    """
    Generate a motor characteristic plot and save it to the specified output 
    inupt directory : need a csv file, loadsort.csv
    """
    with open(input_dir, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers_en = next(reader)
        headers_ch = next(reader)
        # print(f"headers_en: {headers_en}")
        # print(f"headers_ch: {headers_ch}")
        
        csv_data = []
        for row in reader:
            csv_data.append(list(map(float, row)))
            
    plotter = MultiScalePlotter(x_grid_count=5 , y_grid_count=5)
    y_colors = ['red', 'orange', 'green', 'blue', 'purple']
    
    curve_colors = [
        (214, 39, 40),    # 紅
        (255, 127, 14),   # 橘
        (204, 153, 0),    # 暗金黃（黃但不亮）
        (44, 160, 44),    # 綠
        (31, 119, 180)    # 藍
    ]
    y_colors = curve_colors
    plotter.set_x_axis(['轉速(rpm)', '轉差率(%)'])
    # plotter.set_x_axis(['轉速(rpm)'])
    plotter.set_y_axis(['電流 (A )', '轉距 (Nm)', '效率 (% )', '入力 (W )', '出力 (W )'], colors=y_colors)
    
    speed = [row[headers_en.index('speed')] for row in csv_data]
    current = [row[headers_en.index('current')] for row in csv_data]
    torque = [row[headers_en.index('torque')] for row in csv_data]
    eff = [row[headers_en.index('eff')] for row in csv_data]
    power = [row[headers_en.index('power')] for row in csv_data]
    pout = [row[headers_en.index('pout')] for row in csv_data]
    
    plotter.set_curve_data(0, 0, list(zip(speed, current)))
    plotter.set_curve_data(0, 1, list(zip(speed, torque)))
    plotter.set_curve_data(0, 2, list(zip(speed, eff)))
    plotter.set_curve_data(0, 3, list(zip(speed, power)))
    plotter.set_curve_data(0, 4, list(zip(speed, pout)))
    
    # plotter.polt()

    plotter._auto_calculate_all_scales()
    plotter._draw_base()
    plotter._make_tick_labels()

    # plotter.x_tick_labels[1] = ['0.0', '0.2', '0.4', '-0.6', '0.8'] 
    
    # calc the 轉差率
    # (同步轉速 - 轉速) / 同步轉速
    for i in range(len(plotter.x_tick_labels[0])):
        spd = float(plotter.x_tick_labels[0][i])
        if spd != 0:
            slip = (motor.speed - spd) / motor.speed
            plotter.x_tick_labels[1][i] = f"{slip:.2f}"
    
    plotter._draw_ticks()
    plotter._draw_axis_name()
    for y_index, y_name in enumerate(plotter.y_axis_names):
        for x_index, x_name in enumerate(plotter.x_axis_names):
            if plotter.curve_datas[y_index][x_index] is not None:
                color = plotter.y_axis_colors[y_index] if plotter.y_axis_colors else 'black'
                plotter._draw_curve(x_index, y_index, plotter.curve_datas[y_index][x_index], color=color, width=5)

    plotter.save(output_path)
    print(f"[polt_motor_characteristic_fig] Plot saved to {output_path}")


if __name__ == '__main__':
    import json
    file_path = 'test_file\project_file\T123_20250207_223915_hand_merge_cns.motor.json'
    with open(file_path, 'r', encoding='UTF8') as f:
        data = json.load(f)
    
    # Example usage
    motor = Motor()
    motor.from_dict(data)   
    
    input_dir = 'test_file/CSV/2025_0730'
    output_path = 'test_file/gen/電動機特性計算表a.xlsx'
    make_report_a(motor, input_dir, output_path)
    
    output_path = 'test_file/gen/電動機特性計算表b.xlsx'
    make_report_b(motor, input_dir, output_path)
    
    output_path = 'test_file/gen/電動機特性計算表cns.xlsx'
    make_report_cns(motor, input_dir, output_path)
    
    